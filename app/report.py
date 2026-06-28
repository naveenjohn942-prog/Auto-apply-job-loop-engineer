import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_HEADERS = [
    "Rank", "Title", "Company", "Location", "Score",
    "Reasoning", "Skill Matches", "Skill Gaps",
    "Recruiter Email", "Outreach Drafted", "URL",
]
_HEADER_FILL = PatternFill("solid", fgColor="2D6A4F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_COL_WIDTHS = [6, 35, 25, 20, 7, 55, 40, 40, 30, 18, 50]


def build_report(matches: list[dict]) -> bytes:
    """Return XLSX bytes for the quality matches list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"

    ws.append(_HEADERS)
    for cell, width in zip(ws[1], _COL_WIDTHS):
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[cell.column_letter].width = width

    for rank, m in enumerate(matches, start=1):
        ws.append([
            rank,
            m.get("title", ""),
            m.get("company", ""),
            m.get("location", ""),
            m.get("score", ""),
            m.get("reasoning", ""),
            ", ".join(m.get("skill_matches", [])),
            ", ".join(m.get("skill_gaps", [])),
            m.get("recruiter_email") or "—",
            "Yes" if m.get("email_draft") else "No",
            m.get("url", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
